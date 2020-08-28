import csv, io
import openpyxl
from Student.models import *
from Course.models import *
from Auth.models import *
from django.shortcuts import redirect,render
from django.contrib.auth.decorators import user_passes_test
from django.contrib import messages
from django.http import HttpResponseRedirect

#@user_passes_test(lambda u: u.is_staff)
def view_for_staff(request):  
    return render(request,'staff_page.html')

def view_dashboard(request):
    return render(request,'staff_dashboard1.html')

def view_dashboard2(request):
    return render(request,'staff_dashboard2.html')

def course_upload(request):    

    if request.method == "GET":
        return render(request, 'staff_course_upload.html') 
    
    if 'file1' in request.FILES:  
        xlsx_file = request.FILES['file1']     
        if xlsx_file.name.endswith('.xlsx') or xlsx_file.name.endswith('.xls') :
            wb = openpyxl.load_workbook(xlsx_file)
            worksheet = wb[wb.sheetnames[0]]
            headers=[]
            rows = worksheet.iter_rows(min_row=1, max_row=1)
            first_row = next(rows)
            headings = [c.value for c in first_row]
            if headings == ['course_id', 'curriculum', 'course_name', 'course_description', 'credit', 'lecture_hours', 'lab_hours', 'selfstudy_hours']:
                i=1
                for row in worksheet.iter_rows():
                    i=i+1
                    if i<=worksheet.max_row:
                        if not Course.objects.filter(course_id = worksheet["A"+str(i)].value).exists():
                            course = Course.objects.create()
                            course.course_id = worksheet["A"+str(i)].value.upper()
                            curr = Curriculum.objects.get(curri=int(worksheet["B"+str(i)].value))
                            course.curriculum = curr
                            course.course_name = worksheet["C"+str(i)].value.upper()
                            course.course_description = worksheet["D"+str(i)].value
                            course.credit = worksheet["E"+str(i)].value
                            course.lecture_hours = worksheet["F"+str(i)].value
                            course.lab_hours = worksheet["G"+str(i)].value
                            course.selfstudy_hours = worksheet["H"+str(i)].value
                            course.save()        
                messages.success(request, 'Course has been create !')
                return render(request, 'staff_course_upload.html')
            else:
                messages.error(request, 'Invalid Template !')
                return render(request, 'staff_course_upload.html')
        else:
            messages.error(request, 'This file is not .CSV file or Xlxs File!')
            return render(request, 'staff_course_upload.html')
    
    if 'file2' in request.FILES:  
        pre_file = request.FILES['file2'] 
        if pre_file.name.endswith('.xlsx') or pre_file.name.endswith('.xls') :
            wb = openpyxl.load_workbook(pre_file)
            worksheet = wb[wb.sheetnames[0]]
            rows = worksheet.iter_rows(min_row=1, max_row=1)
            first_row = next(rows)
            headings = [c.value for c in first_row]
            if headings == ['Pre', 'Post', 'Group', 'Pre_type']:
                i=1
                for row in worksheet.iter_rows():
                    i=i+1
                    if i<=worksheet.max_row:
                        pre = Prerequisite.objects.create()
                        pre.pre_course = Course.objects.get(course_id=worksheet["A"+str(i)].value) 
                        pre.post_course = Course.objects.get(course_id=worksheet["B"+str(i)].value) 
                        pre.group = int(worksheet["C"+str(i)].value)
                        pre.pre_type = Prerequisite_Type.objects.get(id=int(worksheet["D"+str(i)].value))
                        pre.save()        
                messages.success(request, 'Prerequisite Course has been update !')
                return render(request,'staff_course_upload.html')
            else:
                messages.error(request, 'Invalid Template !')
            return render(request, 'staff_course_upload.html')
        else: 
            messages.error(request, 'This is not .xlxs file !')
            return render(request, 'staff_course_upload.html')
    else : 
        messages.error(request, 'No selected File!')
        return render(request,'staff_course_upload.html')

def view_for_staff(request):  
    return render(request,'staff_page.html')

def create_user(request):    
    template = "staff_create_user.html"
    data = User.objects.all()
    prompt = {
        'profiles': data    
              }

    if request.method == "GET":
        return render(request, template, prompt)    
    
    if request.POST.get('username') and request.POST.get('type'):
        if not User.objects.filter(username = request.POST.get('username')).exists():
            new = User.objects.create_user(request.POST.get('username'), password=request.POST.get('username'))
            if request.POST.get('type')=="1":
                new.is_student = "True"
            if request.POST.get('type')=="2":
                new.is_teacher = "True"
            if request.POST.get('type')=="3":
                new.is_staff = "True"
            new.save()
            messages.success(request, 'Create New Username Successfully')
            return render(request, template)
        else:
            messages.success(request, 'Username is already existed !')
            return render(request, template)
    
    if request.FILES:
        csv_file = request.FILES['file']    
        if csv_file.name.endswith('.csv'):
 
            data_set = csv_file.read().decode('UTF-8')    
            io_string = io.StringIO(data_set)
            reader = csv.reader(io_string, delimiter=',', quotechar='"')
            headers = next(reader, None)
            if headers == ['usrename', 'is_student', 'is_staff', 'is_teacher']:
                for column in csv.reader(io_string, delimiter=',', quotechar="|"):
                    if column:
                        usn = column[0]
                        pw = column[0]
                        if not User.objects.filter(username = usn).exists():
                            Post = User.objects.create_user(usn, password=pw)
                            Post.is_student = column[1]
                            Post.is_staff = column[2]
                            Post.is_teacher = column[3]
                            Post.save()
                messages.success(request, 'Users has been create !')
                return render(request, template)
            else:
                messages.error(request, 'Invalid Template File !')
                return render(request, template)

        if csv_file.name.endswith('.xlsx') or csv_file.name.endswith('.xls') :
            wb = openpyxl.load_workbook(csv_file)
            worksheet = wb[wb.sheetnames[0]]
            rows = worksheet.iter_rows(min_row=1, max_row=1)
            first_row = next(rows)
            headings = [c.value for c in first_row]
            if headings == ['username', 'is_student', 'is_staff', 'is_teacher']:
                i=1
                for row in worksheet.iter_rows():
                    i=i+1
                    if i<=worksheet.max_row:
                        usn=worksheet["A"+str(i)].value
                        pw=worksheet["A"+str(i)].value
                        user = User.objects.filter(username=usn)
                        if user.count()<=0:
                            Post = User.objects.create_user(usn, password=pw)
                            Post.is_student = worksheet["B"+str(i)].value
                            Post.is_staff = worksheet["C"+str(i)].value
                            Post.is_teacher = worksheet["D"+str(i)].value
                            Post.save()
                messages.success(request, 'Users has been created !')
                return render(request,template)
            else:
                messages.error(request, 'Invalid Template !')
                return render(request,template)
        else:
            messages.error(request, 'This file is not a .csv file or .xlsx file !')
            return render(request, template)
    else:
        messages.error(request, 'No file selected')
        return render(request, template)
    
def teacher_upload(request):
    if request.method == "GET":
        return render(request, 'teacher_upload.html') 
        
    if request.FILES:
        xlsx_file =  request.FILES['file']
        if xlsx_file.name.endswith('.xlsx') or xlsx_file.name.endswith('.xls') :
            wb = openpyxl.load_workbook(xlsx_file)
            worksheet = wb[wb.sheetnames[0]]
            rows = worksheet.iter_rows(min_row=1, max_row=1)
            first_row = next(rows)
            headings = [c.value for c in first_row]
            if headings == ['teacher_id', 'academin position', 'education_position', 'name', 'surname', 'email', 'campus', 'status']:
                i=1
                for row in worksheet.iter_rows():
                    i=i+1
                    if i<=worksheet.max_row:
                        user=User.objects.get(username=int(worksheet["A"+str(i)].value))
                        tea = Teacher.objects.filter(teacher_id=user)
                        if tea.count()<=0:
                            teacher = Teacher.objects.create(teacher_id=user)
                            teacher.academic_position = worksheet["B"+str(i)].value
                            if worksheet["C"+str(i)].value != None:
                                teacher.education_position = worksheet["C"+str(i)].value
                            else:
                                teacher.education_position = ""
                            teacher.name = worksheet["D"+str(i)].value
                            teacher.surname = worksheet["E"+str(i)].value
                            teacher.email = worksheet["F"+str(i)].value
                            teacher.campus = worksheet["G"+str(i)].value
                            teacher.status = worksheet["H"+str(i)].value
                            teacher.save()
                        else : 
                            teacher = Teacher.objects.get(teacher_id=user)
                            teacher.academic_position = worksheet["B"+str(i)].value
                            teacher.education_position = worksheet["C"+str(i)].value
                            teacher.name = worksheet["D"+str(i)].value
                            teacher.surname = worksheet["E"+str(i)].value
                            teacher.email = worksheet["F"+str(i)].value
                            teacher.campus = worksheet["G"+str(i)].value
                            teacher.status = worksheet["H"+str(i)].value
                            teacher.save()
                messages.success(request, "Teacher Infomations have been created/updated!")
                return render(request,'teacher_upload.html')
            else:
                messages.success(request, "Invalid Template !")
                return render(request,'teacher_upload.html')

        if xlsx_file.name.endswith('.csv'):
            data_set = xlsx_file.read().decode('UTF-8')    
            io_string = io.StringIO(data_set)
            reader = csv.reader(io_string, delimiter=',', quotechar='"')
            headers = next(reader, None)
            if headers == ['teacher_id', 'academin position', 'education_position', 'name', 'surname', 'email', 'campus', 'status']:
                for column in csv.reader(io_string, delimiter=',', quotechar="|"):
                    if column:
                        usn = column[0]
                        user=User.objects.get(username=usn)
                        tea = Teacher.objects.filter(teacher_id=user)
                        if tea.count()<=0:
                            teacher = Teacher.objects.create(teacher_id = user)
                            teacher.academic_position = column[1]
                            teacher.education_position = column[2]
                            teacher.name = column[3]
                            teacher.surname = column[4]
                            teacher.email = column[5]
                            teacher.campus = column[6]
                            teacher.status = column[7]
                            teacher.save()
                        else : 
                            teacher = Teacher.objects.get(teacher_id=user)
                            teacher.academic_position = column[1]
                            teacher.education_position = column[2]
                            teacher.name = column[3]
                            teacher.surname = column[4]
                            teacher.email = column[5]
                            teacher.campus = column[6]
                            teacher.status = column[7]
                            teacher.save()
                messages.success(request, "Teacher Infomations have been created/updated!")
                return render(request, 'teacher_upload.html')
            else:
                messages.error(request, 'Invalid Template !')
                return render(request, 'teacher_upload.html')
        else:
            messages.error(request, 'This is not .csv file or .xlsx file')
            return render(request, 'teacher_upload.html')
    else :
        messages.error(request, 'No Selected File !')
        return render(request,'teacher_upload.html')

      

    
    if xlsx_file.name.endswith('csv'):
        data_set = xlsx_file.read().decode('UTF-8')    
        io_string = io.StringIO(data_set)
        reader = csv.reader(io_string, delimiter=',', quotechar='"')
        headers = next(reader, None)
        if headers == ['CAMPUSNAME', 'LEVELNAME', 'ADMITACADYEAR', 'APPLICANTTYPEID', 'APPLICANTTYPENAME', 'STUDENTCODE_SALT_HASH', 'ACADYEAR', 'SEMESTER', 'COURSECODE', 'COURSENAMEENG', 'GRADE']:
            for column in csv.reader(io_string, delimiter=',', quotechar="|"):
                if column:
                    user=User.objects.get(username=column[5])
                    stud = Student.objects.filter(student_id=user.id)
                    if stud.count()<=0:
                        stu = Student.objects.create(student_id=user)
                        stu.admit_year = column[2]
                        stu.degree = column[1]
                        stu.campus = column[0]
                        stu.applicant_code = column[3]
                        stu.applicant_type = column[4]
                        stu.save()
                        ent = Enrollment.objects.create(student_id=stu)
                        ent.academic_year = column[6]
                        ent.semester = column[7]
                        ent.grade = column[10]
                        ent.course_id = column[8]
                        ent.course_name = column[9]
                        ent.credit = column[10]
                        ent.save()   
                        last_101=Enrollment.objects.filter(student_id=stu).filter(course_id="CS101").order_by('academic_year','semester').last()
                        if Check_for_lastest.objects.filter(student_id=stu).exists():
                            check_for_cs101 = last_101
                        else:
                            Check_for_lastest.objects.create(student_id=stu,check_for_cs101=last_101)
                        last_223=Enrollment.objects.filter(student_id=stu).filter(course_id="CS223").order_by('academic_year','semester').last()
                        if Check_for_lastest.objects.filter(student_id=stu).exists():
                            check_for_cs223 = last_223
                        else:
                            Check_for_lastest.objects.create(student_id=stu,check_for_cs223=last_enroll)
                        last_300=Enrollment.objects.filter(student_id=stu).filter(course_id="300").order_by('academic_year','semester').last()
                        if Check_for_lastest.objects.filter(student_id=stu).exists():
                            check_for_cs101 = last_300
                        else:
                            Check_for_lastest.objects.create(student_id=stu,check_for_cs300=last_enroll)
                    else:
                        stu = Student.objects.get(student_id=user)
                        ent = Enrollment.objects.create(student_id=stu)
                        ent.academic_year = column[6]
                        ent.semester = column[7]
                        ent.grade = column[10]
                        ent.course_id = column[8]
                        ent.course_name = column[9]
                        ent.credit = column[10]
                        ent.save() 
                        last_101=Enrollment.objects.filter(student_id=stu).filter(course_id="CS101").order_by('academic_year','semester').last()
                        if Check_for_lastest.objects.filter(student_id=stu).exists():
                            check_for_cs101 = last_101
                        else:
                            Check_for_lastest.objects.create(student_id=stu,check_for_cs101=last_101)
                        last_223=Enrollment.objects.filter(student_id=stu).filter(course_id="CS223").order_by('academic_year','semester').last()
                        if Check_for_lastest.objects.filter(student_id=stu).exists():
                            check_for_cs223 = last_223
                        else:
                            Check_for_lastest.objects.create(student_id=stu,check_for_cs223=last_enroll)
                        last_300=Enrollment.objects.filter(student_id=stu).filter(course_id="CS300").order_by('academic_year','semester').last()
                        if Check_for_lastest.objects.filter(student_id=stu).exists():
                            check_for_cs101 = last_300
                        else:
                            Check_for_lastest.objects.create(student_id=stu,check_for_cs300=last_enroll)      
            messages.success(request, 'Enrollment Infomations have been Updated!')
            return render(request,'enrollment_upload.html')
        else:
            messages.error(request, 'Invalid Template!')
            return render(request,'enrollment_upload.html')
            
    else :
        messages.error(request, 'No File Selected !')
        return render(request,'enrollment_upload.html')

def add_student_info(request):
    tea = Teacher.objects.all()
    if request.method == "GET":
        return render(request, 'student_upload.html' , {'advisor' : tea}) 
    
    if request.POST.get("student_id"):   
        if User.objects.filter(username=request.POST.get("student_id")).exists():
            user=User.objects.get(username=request.POST.get("student_id"))
            if not Student.objects.filter(student_id=user).exists():
                advisor = Teacher.objects.get(teacher_id=request.POST.get("advisor"))
                stu = Student.objects.create(student_id=user)
                stu.titles = request.POST.get("title")
                stu.name = request.POST.get("name")
                stu.surname = request.POST.get("surname")
                stu.status = request.POST.get("status")
                stu.degree = request.POST.get("degree")
                stu.admit_year = request.POST.get("admit_year")
                stu.curriculum = request.POST.get("curri")
                stu.major_track = request.POST.get("major")
                stu.campus = request.POST.get("campus")
                stu.adviser = advisor
                stu.save()
                messages.success(request, "Student Infomations have been created!")
                return render(request,'student_upload.html', {'advisor' : tea})
            else:
                messages.success(request, "Student already existed!")
                return render(request,'student_upload.html', {'advisor' : tea})
        else:
            messages.success(request, "Username not found!")
            return render(request,'student_upload.html', {'advisor' : tea})


    if request.FILES:
        xlsx_file =  request.FILES['file']
        if xlsx_file.name.endswith('.xlsx') or xlsx_file.name.endswith('.xls') :
            wb = openpyxl.load_workbook(xlsx_file)
            worksheet = wb[wb.sheetnames[0]]
            rows = worksheet.iter_rows(min_row=1, max_row=1)
            first_row = next(rows)
            headings = [c.value for c in first_row]
            if headings == ['STUDENTCODE_SALT_HASH', 'STUDENT_NAME_SURNAME_SALT_HASH', 'LEVELNAME', 'ADMITACADYEAR', 'APPLICANTTYPEID', 'APPLICANTTYPENAME', 'APPLICANTTYPENAME_MORESPECIFIC', 'STUDENT_HIGHSCHOOL_STUDY_PLAN', 'STUDENT_HIGHSCHOOL', 'STUDENT_HIGHSCHOOL_GPA']:
                i=1
                for row in worksheet.iter_rows():
                    i=i+1
                    if i<=worksheet.max_row:
                        user=User.objects.get(username=(worksheet["A"+str(i)].value))
                        stu = Student.objects.filter(student_id=user)
                        if stu.count()<=0:
                            stu = Student.objects.create(student_id=user)
                            stu.name = worksheet["B"+str(i)].value
                            stu.degree = worksheet["C"+str(i)].value
                            stu.admit_year = worksheet["D"+str(i)].value
                            if stu.admit_year >= 2561: 
                                stu.curri = Curriculum.objects.get(curri=61)
                            stu.applicant_code = int(worksheet["E"+str(i)].value)
                            stu.applicant_type = worksheet["F"+str(i)].value
                            stu.tcas = worksheet["G"+str(i)].value
                            ent_info = Entrance_Info.objects.create()
                            ent_info.high_school_plan = worksheet["H"+str(i)].value
                            ent_info.school = worksheet["I"+str(i)].value
                            if worksheet["J"+str(i)].value != '':
                                ent_info.gpax = worksheet["J"+str(i)].value
                            ent_info.save()
                            stu.ent_info = ent_info
                            stu.save()
                        else : 
                            stu = Student.objects.get(student_id=user)
                            if stu.ent_info != None:
                                ent=Entrance_Info.objects.get(id=stu.ent_info_id)
                                stu.name = worksheet["B"+str(i)].value
                                stu.degree = worksheet["C"+str(i)].value
                                stu.admit_year = worksheet["D"+str(i)].value
                                if stu.admit_year >= 2561: 
                                    stu.curri = Curriculum.objects.get(curri=61)
                                stu.applicant_code = int(worksheet["E"+str(i)].value)
                                stu.applicant_type = worksheet["F"+str(i)].value
                                stu.tcas = worksheet["G"+str(i)].value
                                ent.high_school_plan = worksheet["H"+str(i)].value
                                ent.school = worksheet["I"+str(i)].value
                                if worksheet["J"+str(i)].value != '':
                                    ent.gpax = worksheet["J"+str(i)].value
                                ent.save()
                                stu.save()
                            else:
                                ent_info = Entrance_Info.objects.create()
                                ent_info.high_school_plan = worksheet["H"+str(i)].value
                                ent_info.school = worksheet["I"+str(i)].value
                                if worksheet["J"+str(i)].value != '':
                                    ent.gpax = worksheet["J"+str(i)].value
                                ent_info.save()
                                stu.name = worksheet["B"+str(i)].value
                                stu.degree = worksheet["C"+str(i)].value
                                stu.admit_year = worksheet["D"+str(i)].value
                                if stu.admit_year >= 2561: 
                                    stu.curri = Curriculum.objects.get(curri=61)
                                stu.applicant_code = int(worksheet["E"+str(i)].value)
                                stu.applicant_type = worksheet["F"+str(i)].value
                                stu.tcas = worksheet["G"+str(i)].value
                                stu.ent_info = ent_info
                                stu.save()
                messages.success(request, "Student Infomations have been created/updated!")
                return render(request,'student_upload.html')
            else:
                messages.error(request, "Invalid Template !")
                return render(request,'student_upload.html')

        
        if xlsx_file.name.endswith('.csv'):
            data_set = xlsx_file.read().decode('UTF-8')    
            io_string = io.StringIO(data_set)
            reader = csv.reader(io_string, delimiter=',', quotechar='"')
            headers = next(reader, None)
            
            if headers == ['STUDENTCODE_SALT_HASH', 'STUDENT_NAME_SURNAME_SALT_HASH', 'LEVELNAME', 'ADMITACADYEAR', 'APPLICANTTYPEID', 'APPLICANTTYPENAME', 'APPLICANTTYPENAME_MORESPECIFIC', 'STUDENT_HIGHSCHOOL_STUDY_PLAN', 'STUDENT_HIGHSCHOOL', 'STUDENT_HIGHSCHOOL_GPA']:
                for column in csv.reader(io_string, delimiter=',', quotechar="|"):
                    if column:
                        user=User.objects.get(username=column[0])
                        stu = Student.objects.filter(student_id=user)
                        if stu.count()<=0:
                            stu = Student.objects.create(student_id=user)
                            stu.name = column[1]
                            stu.degree = column[2]
                            stu.admit_year = column[3]
                            stu.applicant_code = column[4]
                            stu.applicant_type = column[5]
                            stu.tcas = column[6]
                            ent_info = Entrance_Info.objects.create()
                            ent_info.high_school_plan = column[7]
                            ent_info.school = column[8]
                            if column[9] != '':
                                ent_info.gpax = column[9]
                            ent_info.save()
                            stu.ent_info = ent_info
                            stu.save()
                        else :  
                            stu = Student.objects.get(student_id=user)
                            if stu.ent_info != None:
                                ent=Entrance_Info.objects.get(id=stu.ent_info_id)
                                stu.name = column[1]
                                stu.degree = column[2]
                                stu.admit_year = column[3]
                                stu.applicant_code = column[4]
                                stu.applicant_type = column[5]
                                stu.tcas = column[6]
                                ent.high_school_plan = column[7]
                                ent.school = column[8]
                                if column[9] != '':
                                    ent.gpax = column[9]
                                ent.save()
                                stu.save()
                            else:
                                ent_info = Entrance_Info.objects.create()
                                ent_info.high_school_plan = column[7]
                                ent_info.school = column[8]
                                if column[9] != '':
                                    ent.gpax = column[9]
                                ent_info.save()
                                stu.name = column[2]
                                stu.degree = column[3]
                                stu.admit_year = column[4]
                                stu.applicant_code = column[5]
                                stu.applicant_type = column[6]
                                stu.tcas = column[7]
                                stu.ent_info = ent_info
                                stu.save()
                messages.success(request, 'Student Infomation has been create !')
                return render(request, 'student_upload.html')
            else:
                messages.error(request, 'Invalid Template !')
                return render(request, 'student_upload.html')

        else:
            messages.success(request, 'This is not .csv file or .xlsx file')
            return render(request, 'student_upload.html')
    else :
        messages.error(request, 'THIS IS NOT A EXEL FILE')
        return render(request,'student_upload.html')

def enrollment_upload(request):
    if request.method == "GET":
        return render(request, 'enrollment_upload.html') 
        
    xlsx_file =  request.FILES['file']
    if xlsx_file.name.endswith('.xlsx') or xlsx_file.name.endswith('.xls') :
        wb = openpyxl.load_workbook(xlsx_file)
        worksheet = wb[wb.sheetnames[0]]
        rows = worksheet.iter_rows(min_row=1, max_row=1)
        first_row = next(rows)
        headings = [c.value for c in first_row]
        if headings == ['STUDENTCODE_SALT_HASH', 'ACADYEAR', 'SEMESTER', 'COURSECODE', 'COURSENAMEENG', 'CREDITSATISFY', 'GRADE']:
            i=1
            for row in worksheet.iter_rows():
                i=i+1
                if i<=worksheet.max_row:
                    user=User.objects.get(username=(worksheet["A"+str(i)].value))
                    stud = Student.objects.filter(student_id=user.id)
                    if stud.count()<=0:
                        stu = Student.objects.create(student_id=user)
                        ent = Enrollment.objects.create(student_id=stu)
                        ent.academic_year = int(worksheet["B"+str(i)].value)
                        ent.semester = int(worksheet["C"+str(i)].value)
                        ent.course_id = worksheet["D"+str(i)].value
                        ent.course_name = worksheet["E"+str(i)].value 
                        ent.credit = worksheet["F"+str(i)].value
                        ent.grade = worksheet["G"+str(i)].value
                        ent.save()
                        last_101=Enrollment.objects.filter(student_id=stu).filter(course_id="CS101").order_by('academic_year','semester').last()
                        #last_grade = last_101.grade
                        if Check_for_lastest.objects.filter(student_id=stu).exists():
                            check_for_cs101 = last_101
                        else:
                            Check_for_lastest.objects.create(student_id=stu,check_for_cs101=last_101)
                        last_223=Enrollment.objects.filter(student_id=stu).filter(course_id="CS223").order_by('academic_year','semester').last()
                        if Check_for_lastest.objects.filter(student_id=stu).exists():
                            check_for_cs223 = last_223
                        else:
                            Check_for_lastest.objects.create(student_id=stu,check_for_cs223=last_enroll)
                        last_300=Enrollment.objects.filter(student_id=stu).filter(course_id="300").order_by('academic_year','semester').last()
                        if Check_for_lastest.objects.filter(student_id=stu).exists():
                            check_for_cs101 = last_300
                        else:
                            Check_for_lastest.objects.create(student_id=stu,check_for_cs300=last_enroll)
                    else:
                        stu = Student.objects.get(student_id=user)
                        ent = Enrollment.objects.create(student_id=stu)
                        ent.academic_year = int(worksheet["B"+str(i)].value)
                        ent.semester = int(worksheet["C"+str(i)].value)
                        ent.course_id = worksheet["D"+str(i)].value
                        ent.course_name = worksheet["E"+str(i)].value 
                        ent.credit = worksheet["F"+str(i)].value
                        ent.grade = worksheet["G"+str(i)].value
                        ent.save()
                        last_101=Enrollment.objects.filter(student_id=stu).filter(course_id="CS101").order_by('academic_year','semester').last()
                        if Check_for_lastest.objects.filter(student_id=stu).exists():
                            check_for_cs101 = last_101
                        else:
                            Check_for_lastest.objects.create(student_id=stu,check_for_cs101=last_101)
                        last_223=Enrollment.objects.filter(student_id=stu).filter(course_id="CS223").order_by('academic_year','semester').last()
                        if Check_for_lastest.objects.filter(student_id=stu).exists():
                            check_for_cs223 = last_223
                        else:
                            Check_for_lastest.objects.create(student_id=stu,check_for_cs223=last_enroll)
                        last_300=Enrollment.objects.filter(student_id=stu).filter(course_id="300").order_by('academic_year','semester').last()
                        if Check_for_lastest.objects.filter(student_id=stu).exists():
                            check_for_cs101 = last_300
                        else:
                            Check_for_lastest.objects.create(student_id=stu,check_for_cs300=last_enroll)
            messages.success(request, "Enrollments have been upload!")
            return render(request,'enrollment_upload.html')
        else:
            messages.error(request, "Invalid Template!")
            return render(request,'enrollment_upload.html')

    
    if xlsx_file.name.endswith('csv'):
        data_set = xlsx_file.read().decode('UTF-8')    
        io_string = io.StringIO(data_set)
        reader = csv.reader(io_string, delimiter=',', quotechar='"')
        headers = next(reader, None)
        if headers == ['CAMPUSNAME', 'LEVELNAME', 'ADMITACADYEAR', 'APPLICANTTYPEID', 'APPLICANTTYPENAME', 'STUDENTCODE_SALT_HASH', 'ACADYEAR', 'SEMESTER', 'COURSECODE', 'COURSENAMEENG', 'GRADE']:
            for column in csv.reader(io_string, delimiter=',', quotechar="|"):
                if column:
                    user=User.objects.get(username=column[5])
                    stud = Student.objects.filter(student_id=user.id)
                    if stud.count()<=0:
                        stu = Student.objects.create(student_id=user)
                        stu.admit_year = column[2]
                        stu.degree = column[1]
                        stu.campus = column[0]
                        stu.applicant_code = column[3]
                        stu.applicant_type = column[4]
                        stu.save()
                        ent = Enrollment.objects.create(student_id=stu)
                        ent.academic_year = column[6]
                        ent.semester = column[7]
                        ent.grade = column[10]
                        ent.course_id = column[8]
                        ent.course_name = column[9]
                        ent.save()   
                        last_101=Enrollment.objects.filter(student_id=stu).filter(course_id="CS101").order_by('academic_year','semester').last()
                        if Check_for_lastest.objects.filter(student_id=stu).exists():
                            check_for_cs101 = last_101
                        else:
                            Check_for_lastest.objects.create(student_id=stu,check_for_cs101=last_101)
                        last_223=Enrollment.objects.filter(student_id=stu).filter(course_id="CS223").order_by('academic_year','semester').last()
                        if Check_for_lastest.objects.filter(student_id=stu).exists():
                            check_for_cs223 = last_223
                        else:
                            Check_for_lastest.objects.create(student_id=stu,check_for_cs223=last_enroll)
                        last_300=Enrollment.objects.filter(student_id=stu).filter(course_id="300").order_by('academic_year','semester').last()
                        if Check_for_lastest.objects.filter(student_id=stu).exists():
                            check_for_cs101 = last_300
                        else:
                            Check_for_lastest.objects.create(student_id=stu,check_for_cs300=last_enroll)
                    else:
                        stu = Student.objects.get(student_id=user)
                        ent = Enrollment.objects.create(student_id=stu)
                        ent.academic_year = column[6]
                        ent.semester = column[7]
                        ent.grade = column[10]
                        ent.course_id = column[8]
                        ent.course_name = column[9]
                        ent.save() 
                        last_101=Enrollment.objects.filter(student_id=stu).filter(course_id="CS101").order_by('academic_year','semester').last()
                        if Check_for_lastest.objects.filter(student_id=stu).exists():
                            check_for_cs101 = last_101
                        else:
                            Check_for_lastest.objects.create(student_id=stu,check_for_cs101=last_101)
                        last_223=Enrollment.objects.filter(student_id=stu).filter(course_id="CS223").order_by('academic_year','semester').last()
                        if Check_for_lastest.objects.filter(student_id=stu).exists():
                            check_for_cs223 = last_223
                        else:
                            Check_for_lastest.objects.create(student_id=stu,check_for_cs223=last_enroll)
                        last_300=Enrollment.objects.filter(student_id=stu).filter(course_id="CS300").order_by('academic_year','semester').last()
                        if Check_for_lastest.objects.filter(student_id=stu).exists():
                            check_for_cs101 = last_300
                        else:
                            Check_for_lastest.objects.create(student_id=stu,check_for_cs300=last_enroll)      
            messages.success(request, 'Enrollment Infomations have been Updated!')
            return render(request,'enrollment_upload.html')
        else:
            messages.error(request, 'Invalid Template!')
            return render(request,'enrollment_upload.html')
            
    else :
        messages.error(request, 'No File Selected !')
        return render(request,'enrollment_upload.html')

def search_student(request):
    advisor = Teacher.objects.all()
    if request.method == "POST":
        if 'search' in request.POST:
            if request.POST.get('stu_id'):
                if User.objects.filter(username=request.POST.get('stu_id')).exists():
                    user = User.objects.get(username=request.POST.get('stu_id'))
                    if Student.objects.filter(student_id=user).exists():
                        stu = Student.objects.get(student_id=user)
                        return render(request,"staff_search_student.html",{'data': stu ,'advisor' : advisor})
                    else : 
                        messages.error(request, 'Student Not Found !')
                        return render(request,"staff_search_student.html",{'advisor' : advisor})
                else:
                    messages.error(request, 'Student Not Found !')
                    return render(request,"staff_search_student.html",{'advisor' : advisor})
            else:
                messages.error(request, 'Please enter Student ID !')
                return render(request,"staff_search_student.html",{'advisor' : advisor})
        if 'update' in request.POST:
            if(request.POST.get('check')!=""):
                user = User.objects.get(username=request.POST.get('check'))
                stu = Student.objects.get(student_id=user)
                if request.POST.get('title'):
                    if request.POST.get('title')!="0":
                        stu.titles= request.POST.get('title')
                if request.POST.get('name'):
                    stu.name= request.POST.get('name')
                if request.POST.get('surname'):
                    stu.surname= request.POST.get('surname')
                if request.POST.get('status'):
                    stu.status= request.POST.get('status')
                if request.POST.get('gpa'):
                    stu.gpa= request.POST.get('gpa')
                if request.POST.get('advisor'): 
                    if request.POST.get('advisor')!="0":
                        stu.adviser= Teacher.objects.get(id=request.POST.get('advisor'))
                if request.POST.get('major'):
                    if request.POST.get('major')!="0":
                         stu.major_track= request.POST.get('major')
                stu.save()
                messages.success(request, 'Update Successfully !')
                return render(request,"staff_search_student.html",{'advisor' : advisor ,'data' : stu}) 
            else:
                messages.error(request, 'Please Select Student  !')
                return render(request,"staff_search_student.html",{'advisor' : advisor}) 
        else:
            return render(request,"staff_search_student.html",{'advisor' : advisor})
    else:
        return render(request,"staff_search_student.html",{'advisor' : advisor})

